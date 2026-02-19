"""
BOM Exporter — Экспорт спецификации в CSV/Excel/PDF
====================================================
УЛУЧШЕНИЯ v4.0:
  - Добавлен PDF экспорт
  - Улучшено форматирование Excel
  - Добавлены итоги и суммы
"""
from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Dict, List

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except Exception:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    PDF_AVAILABLE = True
except Exception:
    PDF_AVAILABLE = False

try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except Exception:
    PANDAS_AVAILABLE = False

from modules.layher_standards import BillOfMaterials, LayherStandards


class BOMExporter:
    """Экспортер спецификации материалов."""

    def export_to_csv(self, bom: BillOfMaterials, project_name: str = "Unnamed") -> str:
        """Экспорт BOM в CSV."""
        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow(["PROJECT", project_name])
        writer.writerow(["DATE", datetime.now().strftime("%Y-%m-%d %H:%M")])
        writer.writerow([])

        writer.writerow(
            [
                "Article Code",
                "Description",
                "Quantity",
                "Unit Weight (kg)",
                "Total Weight (kg)",
                "Unit Price (USD)",
                "Total Price (USD)",
            ]
        )

        for code, count in bom.components.items():
            desc = LayherStandards.ARTICLE_NAMES.get(code, "Unknown")
            unit_weight = LayherStandards.ARTICLE_WEIGHTS.get(code, 0)
            unit_price = LayherStandards.ARTICLE_PRICES.get(code, 0)

            writer.writerow(
                [
                    code,
                    desc,
                    count,
                    f"{unit_weight:.2f}",
                    f"{unit_weight * count:.2f}",
                    f"{unit_price:.2f}",
                    f"{unit_price * count:.2f}",
                ]
            )

        writer.writerow([])
        writer.writerow(
            [
                "",
                "TOTAL",
                bom.get_total_quantity(),
                "",
                f"{bom.get_total_weight():.2f}",
                "",
                f"{bom.get_total_cost():.2f}",
            ]
        )

        return output.getvalue()

    def export_to_excel(
        self, bom: BillOfMaterials, filepath: str, project_name: str = "Unnamed"
    ) -> bool:
        """Экспорт BOM в Excel."""
        if not EXCEL_AVAILABLE:
            print("⚠️  openpyxl not available")
            return False

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BOM"

            ws["A1"] = "BILL OF MATERIALS"
            ws["A1"].font = Font(size=16, bold=True)
            ws["A2"] = f"Project: {project_name}"
            ws["A3"] = f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

            start_row = 5
            headers = [
                "Article Code",
                "Description",
                "Qty",
                "Unit Weight (kg)",
                "Total Weight (kg)",
                "Unit Price (USD)",
                "Total Price (USD)",
            ]

            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

            row = start_row + 1
            for code, count in bom.components.items():
                desc = LayherStandards.ARTICLE_NAMES.get(code, "Unknown")
                unit_weight = LayherStandards.ARTICLE_WEIGHTS.get(code, 0)
                unit_price = LayherStandards.ARTICLE_PRICES.get(code, 0)

                ws.cell(row=row, column=1, value=code)
                ws.cell(row=row, column=2, value=desc)
                ws.cell(row=row, column=3, value=count)
                ws.cell(row=row, column=4, value=unit_weight)
                ws.cell(row=row, column=5, value=unit_weight * count)
                ws.cell(row=row, column=6, value=unit_price)
                ws.cell(row=row, column=7, value=unit_price * count)
                row += 1

            row += 1
            ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
            ws.cell(row=row, column=3, value=bom.get_total_quantity())
            ws.cell(row=row, column=5, value=bom.get_total_weight())
            ws.cell(row=row, column=7, value=bom.get_total_cost())

            for col in range(1, 8):
                ws.column_dimensions[get_column_letter(col)].width = 18

            wb.save(filepath)
            return True

        except Exception as e:
            print(f"Excel export error: {e}")
            return False

    def export_to_pdf(
        self, bom: BillOfMaterials, filepath: str, project_name: str = "Unnamed"
    ) -> bool:
        """Экспорт BOM в PDF."""
        if not PDF_AVAILABLE:
            print("⚠️  reportlab not available")
            return False

        try:
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            story: List = []
            styles = getSampleStyleSheet()

            title_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Heading1"],
                fontSize=18,
                textColor=colors.HexColor("#1f4788"),
                spaceAfter=12,
            )
            story.append(Paragraph("BILL OF MATERIALS", title_style))
            story.append(Spacer(1, 0.3 * cm))

            info_style = styles["Normal"]
            story.append(Paragraph(f"<b>Project:</b> {project_name}", info_style))
            story.append(
                Paragraph(
                    f"<b>Date:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                    info_style,
                )
            )
            story.append(Spacer(1, 0.5 * cm))

            table_data = [
                [
                    "Article",
                    "Description",
                    "Qty",
                    "Weight/unit",
                    "Total Weight",
                    "Price/unit",
                    "Total Price",
                ]
            ]

            for code, count in bom.components.items():
                desc = LayherStandards.ARTICLE_NAMES.get(code, "Unknown")
                unit_weight = LayherStandards.ARTICLE_WEIGHTS.get(code, 0)
                unit_price = LayherStandards.ARTICLE_PRICES.get(code, 0)

                table_data.append(
                    [
                        code,
                        desc,
                        str(count),
                        f"{unit_weight:.1f} kg",
                        f"{unit_weight * count:.1f} kg",
                        f"${unit_price:.2f}",
                        f"${unit_price * count:.2f}",
                    ]
                )

            table_data.append(
                [
                    "",
                    "TOTAL",
                    str(bom.get_total_quantity()),
                    "",
                    f"{bom.get_total_weight():.1f} kg",
                    "",
                    f"${bom.get_total_cost():.2f}",
                ]
            )

            table = Table(table_data, repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, -1), (-1, -1), colors.beige),
                        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ]
                )
            )

            story.append(table)
            doc.build(story)
            return True

        except Exception as e:
            print(f"PDF export error: {e}")
            return False

    def export_to_dataframe(self, bom: BillOfMaterials):
        """Опциональный экспорт BOM в pandas DataFrame (если pandas доступен)."""
        if not PANDAS_AVAILABLE:
            return None

        rows: List[Dict] = []
        for code, count in bom.components.items():
            unit_weight = LayherStandards.ARTICLE_WEIGHTS.get(code, 0)
            unit_price = LayherStandards.ARTICLE_PRICES.get(code, 0)
            rows.append(
                {
                    "article_code": code,
                    "description": LayherStandards.ARTICLE_NAMES.get(code, "Unknown"),
                    "quantity": count,
                    "unit_weight_kg": unit_weight,
                    "total_weight_kg": unit_weight * count,
                    "unit_price_usd": unit_price,
                    "total_price_usd": unit_price * count,
                }
            )

        return pd.DataFrame(rows)


if __name__ == "__main__":
    print("🧪 BOM exporter module loaded. Use from application context for full test.")
